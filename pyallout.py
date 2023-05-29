import cv2  # library cv2 untuk gambar
import serial  # library serial
import time  # library time untuk delay
from openpyxl import Workbook, load_workbook  # library akses excel
from datetime import datetime  # library akses tanggal dan jam

konter = "1"  # variabel konter untuk nomer
# deklarasi serial di COM7 dan baudrate 9600
serialcomm = serial.Serial('COM3', 9600)
serialcomm.timeout = 1  # timeout serial 1ms

face_cascade = cv2.CascadeClassifier(
    'haarcascade_frontalface_default.xml')  # klasifikasi wajah tampak depan
mouth_cascade = cv2.CascadeClassifier(
    'haarcascade_mcs_mouth.xml')  # klasifikasi mulut
cap = cv2.VideoCapture(0)  # baca video dari camera

while True:  # perulangan
    _, img = cap.read()  # baca video
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)  # konversi video ke grayscale
    faces = face_cascade.detectMultiScale(
        gray, 1.2, 6)  # variabel klasifikasi wajah
    mouth = mouth_cascade.detectMultiScale(
        gray, 1.5, 9)  # variabel klasifikasi mulut
    for (x, y, w, h) in faces:  # jika deteksi wajah
        # beri kotak pada daerah yang dideteksi
        cv2.rectangle(img, (x, y), (x+w, y+h), (255, 0, 0), 2)
    for (x, y, w, h) in mouth:  # jika deteksi mulut
        # beri kotak pada daerah yang dideteksi
        cv2.rectangle(img, (x, y), (x+w, y+h), (0, 0, 255), 2)
    font = cv2.FONT_HERSHEY_SIMPLEX  # font untuk teks pada videocapture
    if (len(faces) == 1 and len(mouth) == 0):  # jika terdapat wajah dan tidak terdapat mulut
        cv2.putText(img, 'Wajah', (50, 50), font, 1, (0, 0, 0),
                    2, cv2.LINE_4)  # memberi teks pada video
        serialcomm.write("on".encode())  # mengirim serial "on"
        ketmask = "yes"  # keterangan "yes" pada database
    elif (len(faces) == 1 and len(mouth) == 1):  # jika terdapat wajah dan terdapat mulut
        serialcomm.write("off".encode())  # mengirim serial "off"
        ketmask = "no mask"  # keterangan "no mask" pada database
    elif (len(faces) == 0):  # jika tidak teradapat wajah
        serialcomm.write("off".encode())  # mengirim serial "off"
        ketmask = "nothing"  # keterangan "nothing" pada database

    # membaca serial yang diterima dan di decode dengan ascii
    serialread = serialcomm.readline().decode('ascii')
    now = datetime.now()  # tanggal dan waktu sekarang
    dt_string1 = now.strftime("%d/%m/%Y")  # tanggal dengan format dd/mm/yy
    dt_string2 = now.strftime("%H:%M:%S")  # jam dengan format hh/mm/ss
    wb = load_workbook('cbdb.xlsx')  # buka file excel
    ws = wb['Sheet1']  # nama sheet excel yang dibuka
    konter = int(konter)+1  # konter + 1 setiap perulangan
    ws["A" + str(konter)].value = int(konter) - \
        1  # kolom A pada excel untuk nomer
    # kolom B pada excel untuk tanggal
    ws["B" + str(konter)].value = dt_string1
    ws["C" + str(konter)].value = dt_string2  # kolom C pada excel untuk jam
    # kolom D pada excel untuk suhu dari arduino
    ws["D" + str(konter)].value = serialread
    # kolom E pada excel untuk keterangan
    ws["E" + str(konter)].value = ketmask
    wb.save('cbdb.xlsx')  # save excel

    cv2.imshow('Detected Face', img)  # menampilkan video
    k = cv2.waitKey(30) & 0xff  # jika tekan esc pada keyboard
    if k == 27:
        break  # keluar dari perulangan
    time.sleep(0.5)  # delay 0.5ms
    print(ketmask)  # print keterangan di terminal python

cap.release()  # release video
serialcomm.close()  # menutup serial
